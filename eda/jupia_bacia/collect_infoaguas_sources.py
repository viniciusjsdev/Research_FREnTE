"""Collect CETESB InfoAguas surface-water quality samples for the Jupia report.

Authenticated collection (credentials via INFOAGUAS_EMAIL / INFOAGUAS_SENHA env
vars; never hardcode credentials in the repository). Raw XLSX exports are
written under data/runs (run id prefix ``operational-infoaguas-jupia-``) plus a
consolidated CSV in the run's processing folder, which
process_context_sources.py turns into staging/analytic tables.

Scope: the SP corner of the Jupia contributing basin that descends to the
Parana - UGRHIs 19 (Baixo Tiete), 18 (Sao Jose dos Dourados), 15 (Turvo/
Grande), 12 (Baixo Pardo/Grande) and the "Rio Parana" hydric system - with
the Monitoramento query (all parameters per point) in 5-year windows from 2000.
"""

from __future__ import annotations

import csv
import json
import os
import re
import time
from datetime import datetime, timezone
from html import unescape
from io import BytesIO
from pathlib import Path

import httpx
import openpyxl


ROOT = Path(__file__).resolve().parents[2]
RUNS_DIR = ROOT / "data" / "runs"
BASE = "https://sistemainfoaguas.cetesb.sp.gov.br"
REPORT = f"{BASE}/AguasSuperficiais/RelatorioQualidadeAguasSuperficiais"

FILTERS = [
    ("ugrhi_19_baixo_tiete", {"FiltroTipo": "1", "NUGRHI": "19"}),
    ("ugrhi_18_sj_dourados", {"FiltroTipo": "1", "NUGRHI": "18"}),
    ("ugrhi_15_turvo_grande", {"FiltroTipo": "1", "NUGRHI": "15"}),
    ("ugrhi_12_baixo_pardo_grande", {"FiltroTipo": "1", "NUGRHI": "12"}),
    ("sistema_rio_parana", {"FiltroTipo": "2", "DSISTMHIDRC": "Rio Paraná"}),
]
WINDOW_START_YEAR = 2000
WINDOW_YEARS = 5
REQUEST_PAUSE_S = 0.4


def login(client: httpx.Client) -> None:
    email = os.environ.get("INFOAGUAS_EMAIL", "")
    senha = os.environ.get("INFOAGUAS_SENHA", "")
    if not email or not senha:
        raise SystemExit("Defina INFOAGUAS_EMAIL e INFOAGUAS_SENHA no ambiente.")
    landing = client.get(BASE + "/")
    token = re.search(r'__RequestVerificationToken[^>]*value="([^"]+)"', landing.text).group(1)
    response = client.post(
        BASE + "/",
        data={"__RequestVerificationToken": token, "Email": email, "Senha": senha},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    if '"result":"Ok"' not in response.text:
        raise SystemExit(f"Login InfoAguas falhou: {response.text[:200]}")


def apply_filter(client: httpx.Client, filter_fields: dict[str, str]) -> str:
    page = client.get(REPORT)
    data = {"TipoConsulta": "Monitoramento", **filter_fields}
    token = re.search(r'__RequestVerificationToken[^>]*value="([^"]+)"', page.text)
    if token:
        data["__RequestVerificationToken"] = token.group(1)
    response = client.post(REPORT, data=data, headers={"X-Requested-With": "XMLHttpRequest"})
    return response.text


def list_points(client: httpx.Client) -> list[dict[str, str]]:
    page = client.get(REPORT + "/Monitoramento")
    points: list[dict[str, str]] = []
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", page.text, re.S):
        point_id = re.search(r'name="CodigoPonto"[^>]*value="([^"]+)"', row)
        if not point_id:
            continue
        cells = [unescape(re.sub(r"<[^>]+>|\s+", " ", cell)).strip() for cell in re.findall(r"<td[^>]*>(.*?)</td>", row, re.S)]
        points.append(
            {
                "point_id": point_id.group(1),
                "point_code": cells[0] if cells else "",
                "hydric_system": cells[1] if len(cells) > 1 else "",
                "location": cells[2] if len(cells) > 2 else "",
                "start_date": cells[3] if len(cells) > 3 else "",
                "end_date": cells[4] if len(cells) > 4 else "",
                "municipality": cells[5] if len(cells) > 5 else "",
            }
        )
    return points


def windows_for(start_date: str) -> list[tuple[str, str]]:
    match = re.search(r"(\d{4})$", start_date or "")
    first_year = max(int(match.group(1)) if match else WINDOW_START_YEAR, WINDOW_START_YEAR)
    today = datetime.now()
    windows: list[tuple[str, str]] = []
    year = first_year
    while year <= today.year:
        end_year = min(year + WINDOW_YEARS - 1, today.year)
        end = f"31/12/{end_year}" if end_year < today.year else today.strftime("%d/%m/%Y")
        windows.append((f"01/01/{year}", end))
        year = end_year + 1
    return windows


def fetch_window(client: httpx.Client, point_id: str, start: str, end: str) -> bytes:
    response = client.post(
        REPORT + "/MonitoramentoModal",
        data={"DataInicial": start, "DataFinal": end, "CodigoPonto": point_id},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    response.raise_for_status()
    download = client.get(REPORT + "/Download")
    download.raise_for_status()
    if "spreadsheet" not in download.headers.get("content-type", "") and not download.content[:2] == b"PK":
        raise ValueError(f"Download nao retornou XLSX ({download.headers.get('content-type')})")
    return download.content


def xlsx_rows(content: bytes) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows, [])]
    out = []
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        out.append({h: ("" if c is None else str(c)) for h, c in zip(header, row)})
    workbook.close()
    return out


def main() -> None:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    run_id = f"operational-infoaguas-jupia-{timestamp}"
    run_dir = RUNS_DIR / run_id
    processing_dir = run_dir / "processing"
    processing_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, object]] = []
    all_samples: list[dict[str, str]] = []
    points_catalog: list[dict[str, str]] = []
    seen_points: set[str] = set()

    with httpx.Client(timeout=httpx.Timeout(300, connect=60), follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"}) as client:
        login(client)
        for filter_slug, filter_fields in FILTERS:
            apply_result = apply_filter(client, filter_fields)
            if '"result":"Monitoramento"' not in apply_result:
                results.append({"filter": filter_slug, "status": "error", "notes": apply_result[:200]})
                continue
            points = list_points(client)
            print(f"[{filter_slug}] {len(points)} pontos")
            folder = run_dir / "collection" / "infoaguas_monitoramento" / filter_slug
            folder.mkdir(parents=True, exist_ok=True)
            for point in points:
                if point["point_id"] in seen_points:
                    continue
                seen_points.add(point["point_id"])
                points_catalog.append({**point, "filter": filter_slug})
                for start, end in windows_for(point["start_date"]):
                    slug = f"{point['point_code']}_{start[-4:]}_{end[-4:]}"
                    out_path = folder / f"{slug}.xlsx"
                    try:
                        content = fetch_window(client, point["point_id"], start, end)
                        rows = xlsx_rows(content)
                        if rows:
                            out_path.write_bytes(content)
                        all_samples.extend(rows)
                        results.append({"filter": filter_slug, "point": point["point_code"], "window": f"{start}-{end}", "status": "collected", "rows": len(rows), "bytes": len(content)})
                        print(f"  {slug}: {len(rows)} amostras")
                    except Exception as exc:  # noqa: BLE001 - coleta segue nos demais alvos
                        results.append({"filter": filter_slug, "point": point["point_code"], "window": f"{start}-{end}", "status": "error", "notes": str(exc)[:200]})
                        print(f"  {slug}: ERRO {exc}")
                        # Sessao pode expirar em coletas longas; tenta relogar uma vez.
                        try:
                            login(client)
                            apply_filter(client, filter_fields)
                        except Exception:  # noqa: BLE001
                            pass
                    time.sleep(REQUEST_PAUSE_S)

    samples_path = processing_dir / "infoaguas_samples.csv"
    if all_samples:
        fields = sorted({key for row in all_samples for key in row})
        with samples_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(all_samples)
    catalog_path = processing_dir / "infoaguas_points.csv"
    if points_catalog:
        fields = sorted({key for row in points_catalog for key in row})
        with catalog_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(points_catalog)

    collected = [r for r in results if r.get("status") == "collected"]
    manifest = {
        "run_id": run_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "purpose": "CETESB InfoAguas Monitoramento samples for the SP/Parana corner of the Jupia basin (authenticated user session).",
        "filters": [slug for slug, _ in FILTERS],
        "summary": {
            "targets_total": len(results),
            "collected": len(collected),
            "errors": len(results) - len(collected),
            "points": len(points_catalog),
            "sample_rows": len(all_samples),
        },
        "results": results,
    }
    (run_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest["summary"], indent=2))
    print(f"run: {run_dir}")


if __name__ == "__main__":
    main()
